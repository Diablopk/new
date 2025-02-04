import os
from groq import Groq
import streamlit as st
import pandas as pd
import openpyxl
import unidecode
import json
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import plotly.express as px

# Configuração do tema dark personalizado
st.set_page_config(page_title="Gerador de Planilhas", layout="wide")

# Aplicar tema escuro personalizado
st.markdown("""
    <style>
        /* Tema escuro principal */
        .stApp {
            background: linear-gradient(180deg, #0A0C10 0%, #141621 100%);
            color: #C9D1D9;
        }
        
        /* Sidebar */
        [data-testid="stSidebar"] {
            background: linear-gradient(180deg, #161B22 0%, #1C2128 100%);
            border-right: 1px solid #30363D;
        }
        
        /* Headers */
        h1, h2, h3 {
            color: #58A6FF !important;
            font-weight: 600;
        }
        
        /* Botões */
        .stButton > button {
            background: linear-gradient(180deg, #238636 0%, #2EA043 100%) !important;
            color: white !important;
            border: 1px solid #238636 !important;
            border-radius: 6px !important;
            padding: 0.5rem 1rem !important;
            font-weight: 600 !important;
            transition: all 0.2s !important;
        }
        .stButton > button:hover {
            background: linear-gradient(180deg, #2EA043 0%, #3FB950 100%) !important;
            border-color: #3FB950 !important;
            box-shadow: 0 0 10px rgba(46, 160, 67, 0.4) !important;
        }
        
        /* Inputs e Selectbox */
        .stTextInput > div > div,
        .stSelectbox > div > div,
        .stTextArea > div > div {
            background-color: #0D1117 !important;
            color: #C9D1D9 !important;
            border: 1px solid #30363D !important;
            border-radius: 6px !important;
        }
        
        /* DataFrame */
        .stDataFrame {
            background-color: #0D1117 !important;
            border: 1px solid #30363D !important;
            border-radius: 6px !important;
        }
        .dataframe {
            color: #C9D1D9 !important;
        }
        
        /* Chat messages */
        .stSuccess {
            background-color: #0D1117 !important;
            color: #7CE38B !important;
            border: 1px solid #238636 !important;
            border-radius: 6px !important;
        }
        .stError {
            background-color: #0D1117 !important;
            color: #FF7B72 !important;
            border: 1px solid #F85149 !important;
        }
        .stInfo {
            background-color: #0D1117 !important;
            color: #58A6FF !important;
            border: 1px solid #1F6FEB !important;
        }
        
        /* Expander */
        .streamlit-expanderHeader {
            background-color: #161B22 !important;
            color: #58A6FF !important;
            border: 1px solid #30363D !important;
        }
        
        /* Métricas */
        [data-testid="stMetricValue"] {
            background-color: #161B22 !important;
            color: #7CE38B !important;
            font-weight: 600 !important;
            padding: 1rem !important;
            border-radius: 6px !important;
            border: 1px solid #238636 !important;
        }
        
        /* Separadores */
        hr {
            border-color: #30363D !important;
        }
        
        /* Radio e Checkbox */
        .stRadio > label,
        .stCheckbox > label {
            color: #C9D1D9 !important;
        }
        
        /* Tooltips e hints */
        .stTooltipIcon {
            color: #58A6FF !important;
        }
    </style>
""", unsafe_allow_html=True)

# Atualizar MODELOS_DISPONIVEIS com emojis mais apropriados
MODELOS_DISPONIVEIS = {
    # Modelos Estáveis
    "⚡ LLama 3.3 70B Versatile": "llama-3.3-70b-versatile",
    "🐋 Deepseek-r1": "deepseek-r1-distill-llama-70b",
    "🚀 Mixtral 8x7B": "mixtral-8x7b-32768",
    "💫 Gemma 2 9B": "gemma2-9b-it",
    "⭐ LLama 3.1 8B Instant": "llama-3.1-8b-instant",
    "🌟 LLama 3 70B": "llama3-70b-8192",
    "✨ LLama 3 8B": "llama3-8b-8192",
    
    # Modelos Preview
    "🔮 LLama 3.3 70B SpecDec (Preview)": "llama-3.3-70b-specdec",
    "💎 LLama 3.2 1B (Preview)": "llama-3.2-1b-preview",
    "🌠 LLama 3.2 3B (Preview)": "llama-3.2-3b-preview",
    "✨ LLama 3.2 11B Vision (Preview)": "llama-3.2-11b-vision-preview",
    "⚡ LLama 3.2 90B Vision (Preview)": "llama-3.2-90b-vision-preview"
}

# Configuração da API Groq
GROQ_API_KEY = "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"
client = Groq(api_key=GROQ_API_KEY)

# Configurações iniciais do Streamlit
st.title("Gerador de Planilhas")

# No início do arquivo, após as importações
if 'use_yes_no' not in st.session_state:
    st.session_state.use_yes_no = False

if 'df' not in st.session_state:
    st.session_state.df = None
if 'calculos_aplicados' not in st.session_state:
    st.session_state.calculos_aplicados = False
if 'color_rules' not in st.session_state:
    st.session_state.color_rules = {'valores': {}, 'colunas': {}}

# Adicionar após a inicialização do session_state
if 'edit_rules' not in st.session_state:
    st.session_state.edit_rules = {
        'deleted_cells': set(),  # Armazenar células excluídas (row, col)
        'centered_cells': dict(),  # Modificado para dict para armazenar tipo de alinhamento
        'edited_cells': {}  # Armazenar texto editado {(row, col): novo_texto}
    }

# Adicionar após a inicialização do session_state
if 'headers_edit' not in st.session_state:
    st.session_state.headers_edit = {}

# Adicionar após a inicialização do session_state
if 'chat_messages' not in st.session_state:
    st.session_state.chat_messages = []

# Checkbox com estado persistente
use_yes_no = st.checkbox('Usar formato Sim/Não para listas', 
                        value=st.session_state.use_yes_no,
                        key='use_yes_no')

# Adicionar seletor de modelo
st.sidebar.title("Configurações do Modelo")

# Agrupar modelos por categoria
modelos_estaveis = {k: v for k, v in MODELOS_DISPONIVEIS.items() if "Preview" not in k}
modelos_preview = {k: v for k, v in MODELOS_DISPONIVEIS.items() if "Preview" in k}

categoria_modelo = st.sidebar.radio(
    "Categoria do Modelo:",
    ["Modelos Estáveis", "Modelos Preview"]
)

modelos_disponiveis = modelos_estaveis if categoria_modelo == "Modelos Estáveis" else modelos_preview

modelo_selecionado = st.sidebar.selectbox(
    "Selecione o modelo:",
    list(modelos_disponiveis.keys()),
    index=0
)

# Substituir a seção do chat
st.sidebar.write("---")
st.sidebar.write("### Chat com o Modelo")

with st.sidebar:
    # Área fixa para o chat com altura máxima
    chat_area = st.container()
    chat_area.markdown("""
        <style>
            div.stContainer { max-height: 400px; overflow-y: auto; }
            div.stMarkdown { margin-bottom: 10px; }
        </style>
    """, unsafe_allow_html=True)
    
    with chat_area:
        # Mostrar mensagens
        for msg in st.session_state.chat_messages:
            if msg["role"] == "user":
                st.info(f"**Você**: {msg['content']}")
            else:
                st.success(f"**Assistente**: {msg['content']}")
    
    # Campo de entrada fixo na parte inferior
    with st.container():
        if "temp_input" not in st.session_state:
            st.session_state.temp_input = ""
            
        input_text = st.text_area(
            "Digite sua mensagem:",
            key="user_input",
            height=100
        )
        
        col1, col2 = st.columns([3,1])
        
        with col1:
            if st.button("Enviar", use_container_width=True):
                if input_text.strip():
                    # Adiciona mensagem do usuário
                    st.session_state.chat_messages.append({
                        "role": "user",
                        "content": input_text
                    })
                    
                    try:
                        with st.spinner("Processando..."):
                            response = client.chat.completions.create(
                                model=MODELOS_DISPONIVEIS[modelo_selecionado],
                                messages=[
                                    {"role": "system", "content": "Você deve sempre responder em português do Brasil."},
                                    {"role": "user", "content": input_text}
                                ],
                                temperature=0.7,
                                max_tokens=1000
                            )
                            
                            # Adiciona resposta do assistente
                            st.session_state.chat_messages.append({
                                "role": "assistant",
                                "content": response.choices[0].message.content
                            })
                            
                            # Limpa o campo de entrada
                            st.session_state.temp_input = ""
                            st.rerun()
                            
                    except Exception as e:
                        st.error(f"Erro: {str(e)}")
        
        with col2:
            if st.button("Limpar", use_container_width=True):
                st.session_state.chat_messages = []
                st.session_state.temp_input = ""
                st.rerun()

# Definir funções de cálculo disponíveis
CALCULOS = {
    'Média': lambda x: x.mean(),
    'Soma': lambda x: x.sum(),
    'Máximo': lambda x: x.max(),
    'Mínimo': lambda x: x.min(),
    'Desvio Padrão': lambda x: x.std(),
    'Mediana': lambda x: x.median(),
    'Percentual': lambda x: (x / x.sum()) * 100,
    'Moda': lambda x: x.mode().iloc[0] if not x.mode().empty else None
}

def aplicar_calculos(df, coluna, calculos_selecionados):
    """Aplica os cálculos selecionados na coluna usando apenas nome da função"""
    for calculo in calculos_selecionados:
        if calculo in CALCULOS:
            # Usar apenas o nome do cálculo como nova coluna
            nova_coluna = f"{calculo}"
            df[nova_coluna] = CALCULOS[calculo](df[coluna])
    return df

def process_prompt_to_data(prompt, use_yes_no):
    try:
        formatted_prompt = f"""
        Por favor, responda em português.
        Gere um JSON válido contendo uma lista de objetos para uma tabela.
        O JSON deve ter apenas uma chave principal contendo um array de objetos.
        Exemplo: {{"dados": [{{"coluna1": "valor1", "coluna2": ["item1", "item2"]}}]}}
        Não inclua explicações, apenas o JSON.
        
        Descrição: {prompt}
        """

        response = client.chat.completions.create(
            model=MODELOS_DISPONIVEIS[modelo_selecionado],
            messages=[{"role": "user", "content": formatted_prompt}],
            temperature=0.7,
            max_tokens=2000
        )
        
        content = response.choices[0].message.content
        cleaned_data = clean_json_response(content)
        data = json.loads(cleaned_data)
        
        # Encontrar primeira chave com array
        main_key = next((k for k, v in data.items() if isinstance(v, list)), None)
        
        if (main_key):
            normalized_data = {"usuarios": data[main_key]}
            
            # Aplicar conversão Sim/Não se necessário
            if use_yes_no:
                normalized_data = convert_to_yes_no(normalized_data)
            else:
                normalized_data = normalize_data(normalized_data)
                
            return normalized_data
        else:
            st.error("Formato de dados inválido - necessário array de objetos")
            return None
            
    except Exception as e:
        st.error(f"Erro no processamento: {str(e)}")
        return None

# Função para limpar a resposta JSON
def clean_json_response(response):
    try:
        # Encontrar o JSON mais externo
        start = response.find('{')
        count = 0
        end = -1
        
        if start == -1:
            return '{}'
            
        for i in range(start, len(response)):
            if response[i] == '{':
                count += 1
            elif response[i] == '}':
                count -= 1
                if count == 0:
                    end = i + 1
                    break
                    
        if end == -1:
            return '{}'
            
        json_str = response[start:end]
        
        # Validar JSON
        json.loads(json_str)
        return json_str
        
    except Exception as e:
        st.error(f"Erro na limpeza do JSON: {e}")
        return '{}'

# Função para validar a estrutura do JSON
def validate_json_structure(data):
    try:
        # Validar estrutura esperada
        if not isinstance(data, dict):
            return False
            
        if not any(key in data for key in ['usuarios', 'tabela', 'data']):
            return False
            
        return True
        
    except Exception:
        return False

# Função para normalizar os dados
def normalize_data(data):
    """Normaliza os dados mantendo formatação consistente"""
    try:
        if isinstance(data, dict):
            for key in data.keys():
                if isinstance(data[key], list):
                    normalized_items = []
                    for item in data[key]:
                        normalized_item = {}
                        for k, v in item.items():
                            # Formatar nome da coluna
                            col_name = k.replace('_', ' ').replace('-', ' ').title()
                            
                            if isinstance(v, list):
                                normalized_item[col_name] = ', '.join(str(x).replace('_', ' ').replace('-', ' ') for x in v)
                            else:
                                normalized_item[col_name] = v
                        normalized_items.append(normalized_item)
                    data[key] = normalized_items
        return data
    except Exception as e:
        st.error(f"Erro na normalização: {e}")
        return data

def convert_to_yes_no(data):
    """Converte listas em formato Sim/Não com formatação consistente"""
    if not isinstance(data, dict) or 'usuarios' not in data:
        return data
    
    modified_data = {'usuarios': []}
    
    for item in data['usuarios']:
        new_item = {}
        for k, v in item.items():
            if isinstance(v, list):
                col_prefix = k.replace('_', ' ').replace('-', ' ').title()
                for val in set(v):
                    new_key = f"{col_prefix} {str(val).replace('_', ' ').replace('-', ' ').title()}"
                    new_item[new_key] = 'Sim' if val in v else 'Não'
            else:
                new_item[k.replace('_', ' ').replace('-', ' ').title()] = v
        modified_data['usuarios'].append(new_item)
    
    return modified_data

def format_header(text):
    """Formata o texto substituindo _ e - por espaço"""
    return text.replace('_', ' ').replace('-', ' ').title()

def extract_color_hints(data):
    """Extrai cores especificadas no JSON de resposta"""
    color_hints = {}
    try:
        if isinstance(data, dict) and 'usuarios' in data:
            for item in data['usuarios']:
                for key, value in item.items():
                    if isinstance(value, str):
                        # Busca por padrão valor#COR
                        if '#' in value:
                            val, color = value.split('#', 1)
                            if len(color) == 6 and all(c in '0123456789ABCDEF' for c in color.upper()):
                                val = val.strip()
                                color_hints[val] = color.upper()
    except Exception as e:
        st.error(f"Erro ao extrair cores: {e}")
    return color_hints

def save_data_to_excel(data, filename="relatorio.xlsx"):
    try:
        # Converter dados para DataFrame
        df = pd.DataFrame(data['usuarios'])
        
        # Aplicar edições de cabeçalhos
        if st.session_state.headers_edit:
            df = df.rename(columns=st.session_state.headers_edit)
        
        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Estilos padrão
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        cell_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Adicionar cabeçalhos formatados
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = format_header(column_title)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = cell_border
            cell.alignment = cell_alignment
        
        # Adicionar dados com edições e cores
        for r_idx, row in enumerate(df.values, 2):
            for c_idx, value in enumerate(row, 1):
                # Pular células excluídas
                if (r_idx-2, df.columns[c_idx-1]) in st.session_state.edit_rules['deleted_cells']:
                    continue
                
                cell = ws.cell(row=r_idx, column=c_idx)
                coluna_atual = df.columns[c_idx-1]
                
                # Aplicar texto editado
                if (r_idx-2, coluna_atual) in st.session_state.edit_rules['edited_cells']:
                    cell.value = st.session_state.edit_rules['edited_cells'][(r_idx-2, coluna_atual)]
                else:
                    cell.value = value
                
                # Aplicar alinhamento
                align_type = st.session_state.edit_rules['centered_cells'].get((r_idx-2, coluna_atual), "Esquerda")
                if align_type == "Centro":
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                elif align_type == "Direita":
                    cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                cell.border = cell_border
                
                # Aplicar cores das regras
                valor_str = str(value).strip()
                
                # Verificar cores por valor
                if valor_str in st.session_state.color_rules['valores']:
                    cor = st.session_state.color_rules['valores'][valor_str]
                    cell.fill = PatternFill(start_color=cor, end_color=cor, fill_type='solid')
                # Verificar cores por coluna
                elif coluna_atual in st.session_state.color_rules['colunas']:
                    cor = st.session_state.color_rules['colunas'][coluna_atual]
                    cell.fill = PatternFill(start_color=cor, end_color=cor, fill_type='solid')
                # Manter zebrado para células sem cor específica
                elif r_idx % 2 == 0:
                    cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        # Ajustar largura das colunas
        for col_num in range(1, len(df.columns) + 1):
            col_letter = get_column_letter(col_num)
            max_length = 0
            
            # Verificar comprimento do cabeçalho
            header_cell = ws[f"{col_letter}1"]
            if header_cell.value:
                max_length = len(str(header_cell.value))
            
            # Verificar comprimento dos dados
            for cell in ws[col_letter][1:]:
                if cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            
            # Definir largura da coluna (com margem de 2 caracteres)
            adjusted_width = max_length + 2
            ws.column_dimensions[col_letter].width = min(adjusted_width, 50)  # limita a 50 caracteres
        
        # Ajustar altura das linhas
        for row in ws.rows:
            max_height = 0
            for cell in row:
                if cell.value:
                    text_lines = str(cell.value).count('\n') + 1
                    estimated_height = text_lines * 15  # 15 pontos por linha
                    max_height = max(max_height, estimated_height)
            if max_height > 0:
                ws.row_dimensions[cell.row].height = max_height
        
        wb.save(filename)
        return filename
        
    except Exception as e:
        st.error(f"Erro ao salvar planilha: {e}")
        return None

# Função para extrair colunas numéricas
def extract_numeric_columns(data):
    """Extrai colunas numéricas do DataFrame"""
    if not isinstance(data, pd.DataFrame):
        return None
    
    return data.select_dtypes(include=['int64', 'float64']).columns.tolist()

# Configura interface com Streamlit
st.write("Descreva a planilha que deseja criar e eu cuidarei do resto!")

def analyze_numeric_data(data):
    """Analisa dados numéricos do DataFrame"""
    try:
        df = pd.DataFrame(data['usuarios'])
        numeric_cols = extract_numeric_columns(df)
        
        if not numeric_cols:
            st.warning("Nenhuma coluna numérica encontrada na planilha")
            return
            
        col_to_analyze = st.selectbox(
            "Selecione a coluna para análise",
            numeric_cols
        )
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.write("### Análise Básica")
            st.metric("Soma", f"{df[col_to_analyze].sum():.2f}")
            st.metric("Média", f"{df[col_to_analyze].mean():.2f}")
            st.metric("Máximo", f"{df[col_to_analyze].max():.2f}")
            st.metric("Mínimo", f"{df[col_to_analyze].min():.2f}")
        
        with col2:
            st.write("### Estatísticas")
            st.metric("Mediana", f"{df[col_to_analyze].median():.2f}")
            st.metric("Desvio Padrão", f"{df[col_to_analyze].std():.2f}")
            st.metric("Variância", f"{df[col_to_analyze].var():.2f}")
        
        st.write("### Distribuição")
        fig = px.histogram(df, x=col_to_analyze, title=f"Distribuição de {col_to_analyze}")
        st.plotly_chart(fig, use_container_width=True)
        
    except Exception as e:
        st.error(f"Erro na análise: {str(e)}")

# Substituir a seção do input por:
user_input = st.text_area(
    "Descreva a planilha desejada:",
    placeholder="Exemplo: Crie uma tabela com nomes, idades e cidades",
    height=150  # Define altura do campo em pixels
)

if st.button("Gerar Planilha", key="btn_gerar"):
    if user_input:
        with st.spinner("Processando..."):
            try:
                use_yes_no = st.session_state.use_yes_no
                data = process_prompt_to_data(user_input, use_yes_no)
                
                if data:
                    st.session_state.df = pd.DataFrame(data['usuarios'])
                    st.session_state.calculos_aplicados = False
                    st.success("Planilha gerada com sucesso!")
            except Exception as e:
                st.error(f"Ocorreu um erro: {e}")
    else:
        st.warning("Por favor, digite uma descrição para a planilha")

# Modificar a função que mostra o preview dos dados
def apply_all_edits_to_df(df):
    """Aplica todas as edições ao DataFrame para preview"""
    df_edited = df.copy()
    
    # Aplicar edições de cabeçalhos
    if st.session_state.headers_edit:
        df_edited = df_edited.rename(columns=st.session_state.headers_edit)
    
    # Aplicar edições de texto
    for (idx, col), new_text in st.session_state.edit_rules['edited_cells'].items():
        if col in df_edited.columns and idx < len(df_edited):
            df_edited.loc[idx, col] = new_text
    
    # Formatar números para exibição
    numeric_cols = df_edited.select_dtypes(include=['float64']).columns
    for col in numeric_cols:
        df_edited[col] = df_edited[col].apply(lambda x: f"{x:.1f}")
    
    # Aplicar cores (usando styler)
    def highlight_cells(x):
        styles = pd.DataFrame('', index=x.index, columns=x.columns)
        
        # Aplicar alinhamentos
        for (idx, col), align in st.session_state.edit_rules['centered_cells'].items():
            if col in x.columns and idx < len(x):
                style_text = ""
                if align == "Centro":
                    style_text = "text-align: center;"
                elif align == "Direita":
                    style_text = "text-align: right;"
                elif align == "Esquerda":
                    style_text = "text-align: left;"
                
                if style_text:
                    current_style = styles.loc[idx, col]
                    styles.loc[idx, col] = f"{current_style} {style_text}".strip()
        
        # Cores por valor
        for val, cor in st.session_state.color_rules['valores'].items():
            mask = x.astype(str) == val
            styles = styles.mask(mask, f'background-color: #{cor}')
        
        # Cores por coluna
        for col, cor in st.session_state.color_rules['colunas'].items():
            if col in x.columns:
                styles[col] = f'background-color: #{cor}'
            
        return styles
    
    # Remover células excluídas
    for idx, col in st.session_state.edit_rules['deleted_cells']:
        if col in df_edited.columns and idx < len(df_edited):
            df_edited.loc[idx, col] = None
        
    return df_edited.style.apply(highlight_cells, axis=None)

# Modificar a seção após gerar a planilha
if st.session_state.df is not None:
    # Mostrar preview da planilha original
    st.write("### Planilha Original")
    st.dataframe(st.session_state.df, use_container_width=True)
    
    st.write("---")  # Separador visual
    
    # Seção de edições
    st.write("### Edição da Planilha")
    
    # Adicionar opção para escolher entre dados originais ou com cálculos
    modo_visualizacao = st.radio(
        "Selecione o modo de visualização:",
        ["Dados Originais", "Aplicar Cálculos"],
        key="modo_visualizacao"
    )
    
    df_atual = st.session_state.df.copy()
    
    if modo_visualizacao == "Aplicar Cálculos":
        colunas_numericas = df_atual.select_dtypes(include=['int64', 'float64']).columns
        if len(colunas_numericas) > 0:
            coluna_calculo = st.selectbox("Selecione a coluna para cálculos:", colunas_numericas)
            calculos_selecionados = st.multiselect("Selecione os cálculos:", list(CALCULOS.keys()))
            
            if calculos_selecionados:
                df_atual = aplicar_calculos(df_atual, coluna_calculo, calculos_selecionados)
                st.success("Cálculos aplicados com sucesso!")
        else:
            st.warning("Não há colunas numéricas disponíveis para cálculos")

    # Menu unificado de edição
    st.write("### Ferramentas de Edição")
    tool_menu = st.selectbox(
        "Selecione a ferramenta:",
        ["Formatação de Cores", "Edição de Células", "Edição de Cabeçalhos", "Centralização", "Exclusão de Células"]
    )
    
    with st.expander("Editor de Planilha", expanded=True):
        if tool_menu == "Edição de Cabeçalhos":
            st.write("### Editar Cabeçalhos das Colunas")
            
            # Interface de edição de cabeçalhos
            for col in df_atual.columns:
                col1, col2 = st.columns([3, 1])
                with col1:
                    novo_header = st.text_input(
                        f"Novo nome para '{col}':",
                        value=st.session_state.headers_edit.get(col, col),
                        key=f"header_{col}"
                    )
                with col2:
                    if st.button("Aplicar", key=f"btn_header_{col}"):
                        if novo_header and novo_header != col:
                            st.session_state.headers_edit[col] = novo_header
                            st.success(f"Cabeçalho '{col}' alterado para '{novo_header}'")
                            st.rerun()
            
            if st.button("Limpar Edições de Cabeçalhos"):
                st.session_state.headers_edit = {}
                st.success("Edições de cabeçalhos removidas")
                st.rerun()

        elif tool_menu == "Formatação de Cores":
            # Interface de cores
            modo_cor = st.radio(
                "Modo de coloração:",
                ["Valor específico", "Coluna inteira"]
            )
            
            col1, col2, col3 = st.columns(3)
            with col1:
                coluna_selecionada = st.selectbox("Coluna:", df_atual.columns)
            with col2:
                if modo_cor == "Valor específico":
                    valores_unicos = df_atual[coluna_selecionada].unique()
                    valor_selecionado = st.selectbox("Valor:", [str(v) for v in valores_unicos])
            with col3:
                cor_selecionada = st.color_picker("Cor:", "#FFFFFF")
            
            if st.button("Aplicar Cor"):
                if modo_cor == "Valor específico":
                    st.session_state.color_rules['valores'][str(valor_selecionado)] = cor_selecionada.replace('#', '')
                else:
                    st.session_state.color_rules['colunas'][coluna_selecionada] = cor_selecionada.replace('#', '')
                st.success("Cor aplicada!")
                st.rerun()  # Força atualização do preview

        elif tool_menu == "Edição de Células":
            # Interface de edição
            col1, col2, col3 = st.columns(3)
            with col1:
                edit_col = st.selectbox("Coluna para editar:", df_atual.columns)
            with col2:
                edit_row = st.number_input("Linha (0 = todas):", 0, len(df_atual))
            with col3:
                novo_texto = st.text_input("Novo texto:")
            
            if st.button("Aplicar Edição"):
                if edit_row == 0:
                    for idx in df_atual.index:
                        st.session_state.edit_rules['edited_cells'][(idx, edit_col)] = novo_texto
                else:
                    st.session_state.edit_rules['edited_cells'][(edit_row-1, edit_col)] = novo_texto
                st.success("Edição aplicada!")
                
        elif tool_menu == "Centralização":
            # Interface de alinhamento
            col1, col2, col3 = st.columns(3)
            with col1:
                align_col = st.selectbox("Coluna para alinhar:", df_atual.columns)
            with col2:
                align_row = st.number_input("Linha (0 = todas):", 0, len(df_atual))
            with col3:
                alinhamento = st.selectbox(
                    "Tipo de alinhamento:",
                    ["Esquerda", "Centro", "Direita"]
                )
            
            if st.button("Aplicar Alinhamento"):
                if align_row == 0:
                    for idx in df_atual.index:
                        st.session_state.edit_rules['centered_cells'][(idx, align_col)] = alinhamento
                else:
                    st.session_state.edit_rules['centered_cells'][(align_row-1, align_col)] = alinhamento
                st.success(f"Alinhamento {alinhamento.lower()} aplicado!")
                st.rerun()
                
        elif tool_menu == "Exclusão de Células":
            # Interface de exclusão
            col1, col2 = st.columns(2)
            with col1:
                del_col = st.selectbox("Coluna para excluir:", df_atual.columns)
            
            modo_exclusao = st.radio(
                "Modo de exclusão:",
                ["Excluir linha específica", "Excluir coluna inteira com exceções"]
            )
            
            if modo_exclusao == "Excluir linha específica":
                with col2:
                    del_row = st.number_input("Linha (0 = todas):", 0, len(df_atual))
                
                if st.button("Excluir"):
                    if del_row == 0:
                        for idx in df_atual.index:
                            st.session_state.edit_rules['deleted_cells'].add((idx, del_col))
                    else:
                        st.session_state.edit_rules['deleted_cells'].add((del_row-1, del_col))
                    st.success("Exclusão aplicada!")
            
            else:  # Excluir coluna inteira com exceções
                # Mostrar valores da coluna
                valores_coluna = df_atual[del_col].unique()
                linhas_manter = st.multiselect(
                    "Selecione as linhas que NÃO deseja excluir:",
                    range(len(df_atual)),
                    format_func=lambda x: f"Linha {x+1}: {df_atual.iloc[x][del_col]}"
                )
                
                if st.button("Aplicar Exclusão"):
                    # Excluir toda a coluna exceto as linhas selecionadas
                    for idx in df_atual.index:
                        if idx not in linhas_manter:
                            st.session_state.edit_rules['deleted_cells'].add((idx, del_col))
                    st.success("Exclusão aplicada!")
                    st.rerun()

        # Mostrar preview logo após as ferramentas
        st.write("### Preview dos Dados:")
        df_preview = apply_all_edits_to_df(df_atual)
        st.dataframe(df_preview, use_container_width=True)
    
    # Botões de ação
    col1, col2 = st.columns(2)
    with col1:
        if st.button("Baixar Planilha"):
            filename = save_data_to_excel({'usuarios': df_atual.to_dict('records')})
            if filename:
                with open(filename, "rb") as f:
                    bytes_data = f.read()
                st.download_button(
                    "Download Excel",
                    bytes_data,
                    file_name="planilha_final.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    
    # Corrigir o botão de limpar edições
    with col2:
        if st.button("Limpar Todas as Edições"):
            st.session_state.edit_rules = {
                'deleted_cells': set(),
                'centered_cells': dict(),  # Mudado de set para dict
                'edited_cells': {}
            }
            st.session_state.color_rules = {'valores': {}, 'colunas': {}}
            st.session_state.headers_edit = {}  # Limpar também os cabeçalhos
            st.success("Todas as edições foram removidas!")
            st.rerun()  # Forçar atualização da interface
